import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


def generate_application_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Заявка'

    # Заголовки
    headers = [
        'Фамилия',
        'Имя',
        'Год рождения',
        'Пол (М/Ж)',
        'Дистанция (м)',
        'Стиль',
        'Предварительное время (мм:сс.сс)',
    ]

    # Стиль заголовков
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    header_alignment = Alignment(horizontal='center')

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Ширина колонок
    column_widths = [20, 15, 15, 12, 15, 20, 30]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = width

    # Валидация для пола
    gender_validation = DataValidation(
        type='list',
        formula1='"М,Ж"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle='Неверное значение',
        error='Введите М или Ж',
    )
    ws.add_data_validation(gender_validation)
    gender_validation.add('D2:D1000')

    # Валидация для стиля
    stroke_validation = DataValidation(
        type='list',
        formula1='"кроль,спина,брасс,баттерфляй,комплекс"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle='Неверное значение',
        error='Выберите стиль из списка',
    )
    ws.add_data_validation(stroke_validation)
    stroke_validation.add('F2:F1000')

    # Пример строки
    example_row = [
        'Иванов',
        'Иван',
        2013,
        'М',
        50,
        'кроль',
        '00:32.45',
    ]
    for col, value in enumerate(example_row, start=1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = Font(italic=True, color='999999')

    return wb

def parse_application_excel(file, application):
    import openpyxl
    from apps.teams.models import Athlete, Team

    wb = openpyxl.load_workbook(file)
    ws = wb.active

    stroke_map = {
        'кроль': 'freestyle',
        'спина': 'backstroke',
        'брасс': 'breaststroke',
        'баттерфляй': 'butterfly',
        'комплекс': 'medley',
    }

    gender_map = {
        'м': 'male',
        'ж': 'female',
    }

    errors = []
    athletes_created = []

    for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        # Пропускаем пустые строки
        if not any(row):
            continue

        last_name, first_name, birth_year, gender, distance, stroke, prelim_time = row

        # Валидация обязательных полей
        if not all([last_name, first_name, birth_year, gender, distance, stroke]):
            errors.append(f'Строка {row_num}: не все обязательные поля заполнены')
            continue

        # Нормализация
        gender_normalized = gender_map.get(str(gender).strip().lower())
        if not gender_normalized:
            errors.append(f'Строка {row_num}: неверный пол "{gender}" — используйте М или Ж')
            continue

        stroke_normalized = stroke_map.get(str(stroke).strip().lower())
        if not stroke_normalized:
            errors.append(f'Строка {row_num}: неверный стиль "{stroke}"')
            continue

        try:
            birth_year = int(birth_year)
            distance = int(distance)
        except (ValueError, TypeError):
            errors.append(f'Строка {row_num}: год рождения и дистанция должны быть числами')
            continue

        # Ищем подходящий Event
        from apps.competitions.models import Event, AthleteEvent
        try:
            event = Event.objects.get(
                competition=application.competition,
                gender=gender_normalized,
                stroke=stroke_normalized,
                distance=distance,
                birth_year_from__lte=birth_year,
                birth_year_to__gte=birth_year,
            )
        except Event.DoesNotExist:
            errors.append(
                f'Строка {row_num}: нет категории для '
                f'{gender} {birth_year} {distance}м {stroke}'
            )
            continue

        # Парсим предварительное время
        preliminary_time = None
        if prelim_time:
            try:
                from datetime import timedelta
                parts = str(prelim_time).replace(',', '.').split(':')
                minutes = int(parts[0])
                seconds = float(parts[1])
                preliminary_time = timedelta(
                    minutes=minutes,
                    seconds=int(seconds),
                    milliseconds=int((seconds % 1) * 100)
                )
            except Exception:
                errors.append(f'Строка {row_num}: неверный формат времени "{prelim_time}" — используйте мм:сс.сс')
                continue

        # Создаём спортсмена
        athlete = Athlete.objects.create(
            first_name=str(first_name).strip(),
            last_name=str(last_name).strip(),
            birth_year=birth_year,
            gender=gender_normalized,
            team=application.coach.teams.first(),
            coach=application.coach,
            application=application,
        )

        from apps.competitions.models import AthleteEvent
        AthleteEvent.objects.create(
            athlete=athlete,
            event=event,
            application=application,
            preliminary_time=preliminary_time,
        )

        athletes_created.append(f'{last_name} {first_name}')

    return {
        'success': len(athletes_created),
        'errors': errors,
        'athletes': athletes_created,
    }